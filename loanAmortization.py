import tkinter as tk
from tkinter import messagebox, filedialog, ttk
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
import sqlite3

class LoanCalculator:
    @staticmethod
    def calculate_amortization_schedule(loan_amount, interest_rate, num_payments):
        monthly_rate = interest_rate / 100 / 12
        monthly_payment = loan_amount * (monthly_rate * (1 + monthly_rate) ** num_payments) / ((1 + monthly_rate) ** num_payments - 1)
        
        schedule = []
        balance = loan_amount
        
        for i in range(1, num_payments + 1):
            interest = balance * monthly_rate
            principal = monthly_payment - interest
            balance -= principal
            schedule.append((i, round(monthly_payment, 2), round(interest, 2), round(principal, 2), round(balance, 2) if balance > 0 else 0))
        
        return schedule

class LoanApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Loan Amortization System")
        self.root.geometry("800x650")
        self.root.config(bg="#f4f7fa")

        # Title label with modern style
        title_label = tk.Label(root, text="Loan Amortization Calculator", font=("Helvetica", 20, "bold"), bg="#1a73e8", fg="white", anchor="center")
        title_label.pack(pady=20, fill=tk.X)

        # Main content frame for all input elements with border and padding
        self.main_frame = tk.Frame(root, bg="#f4f7fa", bd=2, relief="solid", padx=20, pady=20)
        self.main_frame.pack(padx=20, pady=20, fill=tk.BOTH, expand=True)

        # Creating input fields with padding and modern style
        self.create_input("Loan Amount:", 0)
        self.create_input("Interest Rate (%):", 1)
        self.create_input("Number of Payments:", 2)

        # Spacer for better separation between input and button areas
        spacer = tk.Label(self.main_frame, text="", bg="#f4f7fa")
        spacer.grid(row=3, column=0, columnspan=2, pady=15)

        # Creating buttons below input fields inside their own frame with border
        self.create_buttons()

    def create_input(self, label, row):
        label_widget = tk.Label(self.main_frame, text=label, font=("Helvetica", 12), bg="#f4f7fa", anchor="w")
        label_widget.grid(row=row, column=0, padx=10, pady=15, sticky="w")
        entry = tk.Entry(self.main_frame, font=("Helvetica", 12), width=25, relief="solid", bd=2, borderwidth=1, highlightthickness=2, highlightcolor="#1a73e8")
        entry.grid(row=row, column=1, padx=10, pady=15, sticky="w", ipadx=5, ipady=5)

        if row == 0: self.entry_loan_amount = entry
        elif row == 1: self.entry_interest_rate = entry
        elif row == 2: self.entry_num_payments = entry

    def create_buttons(self):
        # Frame for the buttons with clear separation from input area
        button_frame = tk.Frame(self.main_frame, bg="#f4f7fa")
        button_frame.grid(row=4, column=0, columnspan=2, pady=15)

        self.create_button(button_frame, "Calculate Loan", self.calculate_loan)
        self.create_button(button_frame, "Clear Inputs", self.clear_inputs)
        self.create_button(button_frame, "Export to Excel", self.export_to_excel)
        self.create_button(button_frame, "Load from Database", self.load_from_database)

    def create_button(self, parent, text, command):
        button = tk.Button(parent, text=text, font=("Helvetica", 12, "bold"), bg="#1a73e8", fg="white", relief="solid", bd=1, padx=10, pady=10, command=command, activebackground="#155a8a", activeforeground="white", borderwidth=2)
        button.pack(side=tk.LEFT, padx=15)

    def clear_inputs(self):
        self.entry_loan_amount.delete(0, tk.END)
        self.entry_interest_rate.delete(0, tk.END)
        self.entry_num_payments.delete(0, tk.END)

    def calculate_loan(self):
        try:
            loan_amount = float(self.entry_loan_amount.get())
            interest_rate = float(self.entry_interest_rate.get())
            num_payments = int(self.entry_num_payments.get())
            self.schedule = LoanCalculator.calculate_amortization_schedule(loan_amount, interest_rate, num_payments)
            self.display_amortization_schedule(self.schedule)
        except ValueError:
            messagebox.showerror("Input Error", "Please enter valid numeric values.")

    def load_from_database(self):
        # Open file dialog to choose database
        db_file = filedialog.askopenfilename(filetypes=[("SQLite Database", "*.sqlite;*.db"), ("All Files", "*.*")])
        
        if db_file:
            try:
                conn = sqlite3.connect(db_file)
                cursor = conn.cursor()

                # Example query: adjust to match your database schema
                cursor.execute("SELECT loan_amount, interest_rate, num_payments FROM loans WHERE id = 1")
                result = cursor.fetchone()

                if result:
                    loan_amount, interest_rate, num_payments = result
                    self.entry_loan_amount.delete(0, tk.END)
                    self.entry_loan_amount.insert(0, str(loan_amount))
                    self.entry_interest_rate.delete(0, tk.END)
                    self.entry_interest_rate.insert(0, str(interest_rate))
                    self.entry_num_payments.delete(0, tk.END)
                    self.entry_num_payments.insert(0, str(num_payments))
                else:
                    messagebox.showerror("Database Error", "No data found for the loan.")
                
                conn.close()
            except sqlite3.Error as e:
                messagebox.showerror("Database Error", f"Failed to load data from database: {e}")

    def display_amortization_schedule(self, schedule):
        window = tk.Toplevel(self.root)
        window.title("Amortization Table")
        window.geometry("750x450")
        window.config(bg="#f4f7fa")

        style = ttk.Style()
        style.configure("Treeview.Heading", font=("Helvetica", 12, "bold"))
        style.configure("Treeview", font=("Helvetica", 10), rowheight=30, highlightthickness=1, highlightcolor="#1a73e8")

        columns = ("Payment", "Prestation", "Interest", "Amortization", "Balance")
        tree = ttk.Treeview(window, columns=columns, show="headings", style="Treeview")
        
        for col in columns:
            tree.heading(col, text=col)
            tree.column(col, anchor="center", width=120)

        tree.tag_configure('evenrow', background='#e9ecef')
        tree.tag_configure('oddrow', background='#f0f4f8')

        for i, row in enumerate(schedule):
            tag = 'evenrow' if i % 2 == 0 else 'oddrow'
            tree.insert("", "end", values=row, tags=(tag,))
        
        tree.pack(fill=tk.BOTH, expand=True)

    def export_to_excel(self):
        if not hasattr(self, 'schedule'):
            messagebox.showwarning("Warning", "Please calculate loan schedule first.")
            return

        filepath = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])
        if filepath:
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Amortization Schedule"

            headers = ["Payment", "Prestation", "Interest", "Amortization", "Balance"]
            for col_num, header in enumerate(headers, 1):
                cell = sheet.cell(row=1, column=col_num, value=header)
                cell.font = Font(bold=True, size=12)
                cell.alignment = Alignment(horizontal="center")
                cell.border = Border(bottom=Side(style="thin"))
                cell.fill = PatternFill(start_color="1a73e8", end_color="1a73e8", fill_type="solid")

            for row_idx, row_data in enumerate(self.schedule, start=2):
                for col_idx, value in enumerate(row_data, start=1):
                    cell = sheet.cell(row=row_idx, column=col_idx, value=value)
                    cell.font = Font(size=10)
                    cell.alignment = Alignment(horizontal="center")
                    cell.border = Border(
                        left=Side(style="thin"),
                        right=Side(style="thin"),
                        top=Side(style="thin"),
                        bottom=Side(style="thin")
                    )

            for col in sheet.columns:
                max_length = max(len(str(cell.value)) for cell in col)
                sheet.column_dimensions[col[0].column_letter].width = max_length + 2

            workbook.save(filepath)
            messagebox.showinfo("Export Successful", "Amortization schedule exported successfully!")

if __name__ == "__main__":
    root = tk.Tk()
    app = LoanApp(root)
    root.mainloop()
